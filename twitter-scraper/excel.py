from openpyxl import Workbook
import openpyxl as xl

filename ="C:\\Users\\Srijib\\.vscode\\reddit-work\\twitter-scraper\\s.xlsx"
wb1 = xl.load_workbook(filename)
ws1 = wb1.worksheets[0]

keywords = ["scanxiety", "scananxiety"]
spcl = ["scan anxiety", "scan-anxiety", "scan-related anxiety", "scan-associated anxiety"]

mr = ws1.max_row

for i in range (3, mr + 1):
    # HYPERLINK
    # c = ws1.cell(row = i, column = 5).value
    # cell = 'E'+str(i)
    # ws1[cell].hyperlink = c
    # ws1[cell].value = "X"

    # REPLY
    # if ws1.cell(row = i, column = 9).value[2] == '@':
    #     ws1.cell(row = i, column = 10).value = 'R'

    # REMOVE b'
    # ws1.cell(row = i, column = 9).value = ws1.cell(row = i, column = 9).value[2:-1]
    # ws1.cell(row = i, column = 11).value = ws1.cell(row = i, column = 11).value[2:-1]
    
    # REMOVE COMMON STUFF
    # v = str(ws1.cell(row = i, column = 13).value)
    # print(v)
    # if v:
    #     vl = v.split(',')
    #     print(vl)
    #     new_vl = ''
    #     for keyword in keywords:
    #         for hashy in vl:
    #             if keyword in hashy:
    #                 new_vl = keyword
    #     ws1.cell(row = i, column = 13).value = new_vl

    # content = ws1.cell(row=i, column=1).value
    # if content:
    #     c = content.split()
    #     hashtag = ''
    #     key = ''
    #     for word in c:
    #         for keyword in keywords:
    #             if keyword in word.lower():
    #                 if word[0] == '#':
    #                     hashtag = hashtag + keyword + ','
    #                 else:
    #                     key = key + keyword + ','
    #     for sp in spcl:
    #         if sp in content:
    #             key = key + sp + ','
    #     if key:
    #         key = key[:-1]
    #     if hashtag:
    #         hashtag = hashtag[:-1]
    #     ws1.cell(row=i, column=2).value = key
    #     ws1.cell(row=i, column=3).value = hashtag
    pass
    
wb1.save(filename=filename)