from openpyxl import Workbook
import openpyxl as xl

filename ="C:\\Users\\Srijib\\.vscode\\reddit-work\\excel-copy\\May 2020 loads.xlsx"
wb1 = xl.load_workbook(filename) 
ws1 = wb1.worksheets[0] 

wb2 = Workbook()

#CHANGE
days = ["Wednesday", "Thursday", "Friday", "Saturday", "Sunday", "Monday", "Tuesday"]

#CHANGE
for date in range(1,32):
    day = days[(date-1)%7]
    if day == "Saturday" or day == "Sunday":
        continue
    
    ws2 = wb2.create_sheet()
    
    #CHANGE
    if date < 10:
        ws2.title = day + " 120" + str(date) + "21"
    else:
        ws2.title = day + " 12" + str(date) + "21"
        

    # Calculate total number of rows and  columns in source excel file 
    mr = ws1.max_row 
    mc = ws1.max_column 

    from copy import copy

    # copying the cell values from source  
    # excel file to destination excel file 
    for i in range (1, mr + 1): 
        for j in range (1, mc + 1): 
            # reading cell value from source excel file 
            c = ws1.cell(row = i, column = j) 
    
            # writing the read value to destination excel file 
            ws2.cell(row = i, column = j).value = c.value 

            if c.has_style:
                ws2.cell(row = i, column = j).font = copy(c.font)
                ws2.cell(row = i, column = j).border = copy(c.border)
                ws2.cell(row = i, column = j).fill = copy(c.fill)
                ws2.cell(row = i, column = j).number_format = copy(c.number_format)
                ws2.cell(row = i, column = j).protection = copy(c.protection)
                ws2.cell(row = i, column = j).alignment = copy(c.alignment)

    for idx, rd in ws1.row_dimensions.items():
        ws2.row_dimensions[idx] = copy(rd)

    for idx, cd in ws1.column_dimensions.items():
        ws2.column_dimensions[idx] = copy(cd)

    # CHANGE
    if date < 10:
        ws2.cell(row = 1, column = 1).value = day + " 12-0" + str(date) + "-21"
    else:
        ws2.cell(row = 1, column = 1).value = day + " 12-" + str(date) + "-21"

    ws2.merge_cells('A1:K1') 
    ws2.merge_cells('A2:K2') 
    ws2.merge_cells('A19:K19') 
    ws2.merge_cells('A25:K25') 

    if day == 'Monday':
        ws2.sheet_properties.tabColor = 'A13D2D'
    elif day == 'Tuesday':
        ws2.sheet_properties.tabColor = 'FF0000'
    elif day == 'Wednesday':
        ws2.sheet_properties.tabColor = 'D5B60A'
    elif day == 'Thursday':
        ws2.sheet_properties.tabColor = '00008B'
    elif day == 'Friday':
        ws2.sheet_properties.tabColor = 'FFC0CB'
    else:
        w2s.sheet_properties.tabColor = '000000'

#CHANGE
wb2.save(filename = './excel-copy/December 2021 Loads.xlsx')